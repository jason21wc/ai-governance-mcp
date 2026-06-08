"""Spreadsheet connector for parsing CSV and Excel files.

Extracts schema (column names, types) and summary statistics.
Indexes headers and sample rows for discoverability.
"""

import csv
import logging
from pathlib import Path

from .base import BaseConnector
from ..models import ContentChunk, FileMetadata

logger = logging.getLogger("ai_governance_mcp.context_engine.connectors.spreadsheet")


class SpreadsheetConnector(BaseConnector):
    """Connector for CSV and Excel files."""

    def __init__(self) -> None:
        self._openpyxl_available = False
        try:
            import openpyxl  # noqa: F401

            self._openpyxl_available = True
        except ImportError:
            pass

    @property
    def supported_extensions(self) -> set[str]:
        extensions = {".csv", ".tsv"}
        if self._openpyxl_available:
            extensions.add(".xlsx")
        return extensions

    def can_handle(self, file_path: Path) -> bool:
        return file_path.suffix.lower() in self.supported_extensions

    def parse(
        self, file_path: Path, project_root: Path | None = None
    ) -> list[ContentChunk]:
        """Parse a spreadsheet into schema + sample chunks."""
        # Compute display path (relative to project root when available)
        if project_root and file_path.is_relative_to(project_root):
            display_path = str(file_path.relative_to(project_root))
        else:
            display_path = str(file_path)

        suffix = file_path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            return self._parse_csv(file_path, suffix, display_path)
        elif suffix == ".xlsx":
            return self._parse_xlsx(file_path, display_path)
        return []

    def extract_metadata(self, file_path: Path) -> FileMetadata:
        stat = file_path.stat()
        return FileMetadata(
            path=str(file_path),
            content_type="data",
            language=file_path.suffix.lstrip("."),
            size_bytes=stat.st_size,
            last_modified=stat.st_mtime,
        )

    def _parse_csv(
        self, file_path: Path, suffix: str, display_path: str
    ) -> list[ContentChunk]:
        """Parse CSV/TSV files."""
        chunks: list[ContentChunk] = []
        delimiter = "\t" if suffix == ".tsv" else ","

        try:
            with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f, delimiter=delimiter)
                rows = []
                for i, row in enumerate(reader):
                    # Limit columns to prevent memory exhaustion on wide CSVs
                    rows.append(row[:500])
                    if i >= 10:  # Header + 10 sample rows
                        break

            if not rows:
                return []

            # Schema chunk
            headers = rows[0] if rows else []
            schema_text = f"Schema: {', '.join(headers)}\n"
            schema_text += f"Columns: {len(headers)}\n"
            if len(rows) > 1:
                schema_text += f"Sample rows ({min(len(rows) - 1, 10)}):\n"
                for row in rows[1:]:
                    schema_text += f"  {', '.join(row)}\n"

            chunks.append(
                ContentChunk(
                    content=schema_text,
                    source_path=display_path,
                    start_line=1,
                    end_line=len(rows),
                    content_type="data",
                    heading=f"Schema: {file_path.name}",
                )
            )
        except Exception as e:
            logger.warning("Failed to parse CSV %s: %s", file_path, e)
            return []

        return chunks

    def _parse_xlsx(self, file_path: Path, display_path: str) -> list[ContentChunk]:
        """Parse Excel files."""
        if not self._openpyxl_available:
            return []

        chunks: list[ContentChunk] = []

        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        rows.append(
                            [
                                str(cell) if cell is not None else ""
                                for cell in row[:500]
                            ]
                        )
                        if i >= 10:
                            break

                    if not rows:
                        continue

                    headers = rows[0]
                    schema_text = f"Sheet: {sheet_name}\n"
                    schema_text += f"Schema: {', '.join(headers)}\n"
                    schema_text += f"Columns: {len(headers)}\n"
                    if len(rows) > 1:
                        schema_text += f"Sample rows ({min(len(rows) - 1, 10)}):\n"
                        for row in rows[1:]:
                            schema_text += f"  {', '.join(row)}\n"

                    chunks.append(
                        ContentChunk(
                            content=schema_text,
                            source_path=display_path,
                            start_line=1,
                            end_line=len(rows),
                            content_type="data",
                            heading=f"Schema: {file_path.name} / {sheet_name}",
                        )
                    )
            finally:
                wb.close()
        except Exception as e:
            logger.warning("Failed to parse Excel %s: %s", file_path, e)
            return []

        return chunks
